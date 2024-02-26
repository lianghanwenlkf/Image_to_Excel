import csv
import os
import requests
import yaml
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont


def read_yaml_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        data = yaml.safe_load(file)
    return data


def download_wechat_head_img(wechat_path, head_img_path):
    with open(wechat_path, 'r', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file)

        for i, row in enumerate(csv_reader):
            if i > 0:
                if row[3]:
                    name = row[3]
                else:
                    name = row[4]
                url = row[7]
                try:
                    response = requests.get(url)

                    # 如果请求成功 (状态码为200)
                    if response.status_code == 200:
                        # 以二进制模式写入文件
                        head_img_file_path = head_img_path + '/' + name + '.jpg'
                        with open(head_img_file_path, 'wb') as file:
                            file.write(response.content)
                        print(f"{i}, 文件下载成功: {head_img_file_path}")
                    else:
                        print(f"{i}, 文件下载失败, 状态码: {response.status_code}")

                except Exception as e:
                    print(f"发生错误: {e}")


def add_chinese_text(input_image_path, output_image_path, chinese_text, configs):
    # 打开原始图片
    original_image = Image.open(input_image_path)

    # 获取原始图片宽度和高度
    image_width, image_height = original_image.size

    # 创建空白画布，宽度为原始图片的两倍，高度相同
    canvas = Image.new('RGB', (image_width * 2, image_height), color=(255, 255, 255))

    # 在画布上绘制原始图片
    canvas.paste(original_image, (0, 0))

    # 创建绘图对象
    draw = ImageDraw.Draw(canvas)

    # 设置中文文字相关参数（字体、大小、颜色等）
    font_path = configs["font_path"]
    font_size = configs["font_size"]
    font_color = tuple(configs["font_color"])

    # 加载中文字体
    chinese_font = ImageFont.truetype(font_path, font_size)

    # 在画布上右侧插入中文文字
    text_position = (image_width, image_height // 2)  # 右侧中间位置
    draw.text(text_position, chinese_text, font=chinese_font, fill=font_color)

    # 保存修改后的图片
    canvas.save(output_image_path)


def read_image_and_write_to_excel(image_path, excel_path):
    # 打开图片
    image = Image.open(image_path)

    # 获取图片的宽度和高度
    width, height = image.size

    # 创建一个新的Excel工作簿
    workbook = Workbook()
    sheet = workbook.active

    for y in range(height):
        sheet.row_dimensions[3 * y + 1].height = 15
        sheet.row_dimensions[3 * y + 2].height = 15
        sheet.row_dimensions[3 * y + 3].height = 15
        for x in range(width):
            # 获取像素的RGB颜色值
            pixel_color = image.getpixel((x, y))

            # 将颜色值写入Excel单元格
            cell = sheet.cell(row=3 * y + 1, column=x+1)
            cell.value = pixel_color[0]
            cell.fill = get_gradient_fill(pixel_color[0], 'R')

            cell = sheet.cell(row=3 * y + 2, column=x + 1)
            cell.value = pixel_color[1]
            cell.fill = get_gradient_fill(pixel_color[1], 'G')

            cell = sheet.cell(row=3 * y + 3, column=x + 1)
            cell.value = pixel_color[2]
            cell.fill = get_gradient_fill(pixel_color[2], 'B')

    # 保存Excel文件
    workbook.save(excel_path)


def get_gradient_fill(value, channel):
    # 将值映射到 0 到 1 的范围
    normalized_value = value / 255.0

    # 使用渐变色，根据通道设置颜色
    if channel == 'R':
        fill_color = f'{int(normalized_value * 255):02X}0000'
    elif channel == 'G':
        fill_color = f'00{int(normalized_value * 255):02X}00'
    else:
        fill_color = f'0000{int(normalized_value * 255):02X}'

    # 创建 PatternFill 对象
    pattern_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    return pattern_fill


def list_files_in_directory(folder_path):
    # 检查文件夹路径是否存在
    if not os.path.exists(folder_path):
        print(f"The folder '{folder_path}' does not exist.")
        return

    # 获取文件夹内所有文件的名称
    files = [f for f in os.listdir(folder_path) if f.lower().endswith('.jpg') and os.path.isfile(os.path.join(folder_path, f))]

    return files


def create_folder(folder_path):
    # 检查文件夹是否存在
    if not os.path.exists(folder_path):
        # 如果不存在，则创建文件夹
        os.makedirs(folder_path)


def main():
    # 读取参数
    yaml_file_path = 'configs.yaml'
    configs = read_yaml_file(yaml_file_path)

    # 原始数据路径
    wechat_path = configs['wechat_path']

    # 文件夹路径
    head_img_path = configs['head_img_path']
    text_img_path = configs['text_img_path']
    excel_save_path = configs['excel_save_path']
    create_folder(head_img_path)
    create_folder(text_img_path)
    create_folder(excel_save_path)

    if configs['download_head_img_mode']:
        download_wechat_head_img(wechat_path, head_img_path)

    if configs['add_text_mode']:
        files_list = list_files_in_directory(head_img_path)
        for file in tqdm(files_list):
            input_image_path = head_img_path + '/' + file
            output_image_path = text_img_path + '/' + file
            name = file[:-4]
            chinese_text = f'祝{name}\n{configs["add_text"]}'

            add_chinese_text(input_image_path, output_image_path, chinese_text, configs)

    if configs['img_to_excel_mode']:
        files_list = list_files_in_directory(text_img_path)
        for file in tqdm(files_list):
            name = file[:-4]
            input_image_path = text_img_path + '/' + file
            output_excel_path = excel_save_path + '/' + name
            create_folder(output_excel_path)
            output_excel_file_path = output_excel_path + f'/{configs["output_file_name"]}.xlsx'

            read_image_and_write_to_excel(input_image_path, output_excel_file_path)


if __name__ == '__main__':
    main()
